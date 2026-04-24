from openpyxl import load_workbook
from io import BytesIO


class InvalidImportSchemaError(Exception):
    pass


def parse_transactions_from_file(file_bytes: bytes) -> tuple[list[dict], int]:
    """
    Parses the transactions from the Excel file.
    Returns:
    - a list of parsed transactions dictionaries, each containing the transaction data.
    - the number of skipped invalid rows.
    """
    relevant_headers = [
        "תאריך עסקה",
        "שם בית העסק",
        "קטגוריה",
        "סכום חיוב",
        "מטבע חיוב",
    ]
    transactions = []
    skipped_rows = 0

    try:
        workbook = load_workbook(filename=BytesIO(file_bytes))
    except Exception:
        raise InvalidImportSchemaError("The uploaded file is not a valid Excel file")

    try:
        current_transactions_sheet = workbook["עסקאות במועד החיוב"]
    except KeyError:
        raise InvalidImportSchemaError("The transactions sheet does not exist with the expected name")

    headers = [cell.value for cell in current_transactions_sheet[4]]

    try:
        relevant_header_indices = {
            header: headers.index(header) for header in relevant_headers
        }
    except ValueError:
        raise InvalidImportSchemaError("File is missing one of the expected headers.")

    for row in current_transactions_sheet.iter_rows(min_row=5, values_only=True):
        if not any(row):
            continue

        merchant_name = row[relevant_header_indices["שם בית העסק"]]
        transaction_date = row[relevant_header_indices["תאריך עסקה"]]
        category = row[relevant_header_indices["קטגוריה"]]
        amount = row[relevant_header_indices["סכום חיוב"]]
        currency = row[relevant_header_indices["מטבע חיוב"]]

        merchant_name = str(merchant_name).strip() if merchant_name else None
        category = str(category).strip() if category else None
        currency = str(currency).strip() if currency else None

        if not merchant_name or not transaction_date or amount is None:
            print("Skipping bad row:", row)
            skipped_rows += 1
            continue

        transaction = {
            "transaction_date": transaction_date,
            "merchant_name": merchant_name,
            "category": category,
            "amount": amount,
            "currency": currency,
        }

        transactions.append(transaction)

    return transactions, skipped_rows
